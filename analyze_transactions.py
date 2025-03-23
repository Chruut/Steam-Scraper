import csv
import re
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

def parse_amount(amount_str):
    """Extracts the numeric value from an amount like 'CHF 2.30' or '+CHF 2.30'"""
    if not amount_str:
        return 0.0
        
    # Remove all whitespace, line breaks and tabs
    amount_str = re.sub(r'\s+', '', amount_str)
    
    # Remove 'CHF' and replace comma with dot
    amount_str = amount_str.replace('CHF', '').replace(',', '.')
    
    # Remove '--' and replace with '00'
    amount_str = amount_str.replace('--', '00')
    
    # Remove all non-numeric characters except dot and minus
    amount_str = re.sub(r'[^\d.-]', '', amount_str)
    
    try:
        return float(amount_str)
    except ValueError:
        print(f"Could not parse amount: '{amount_str}'")
        return 0.0

def is_expense(description, change_amount):
    """Checks if it's an expense"""
    # If the change amount is negative, it's an expense
    return change_amount < 0

def is_external_payment(description, payment_info):
    """Checks if it's an external payment (PayPal, credit card)"""
    # Check for external payment methods
    if "PayPal" in payment_info or "MasterCard" in payment_info:
        return True
    # Check for Wallet Credit purchases
    if "Wallet Credit" in description:
        return True
    return False

def is_game_purchase(description):
    """Checks if it's a real game purchase"""
    # Ignore Market transactions
    if "Steam Community Market" in description:
        return False
    # Ignore Refunds
    if "Refund" in description:
        return False
    # Only consider real purchases
    return "Purchase" in description or "In-Game Purchase" in description

def is_market_transaction(description):
    """Checks if it's a Market transaction"""
    return "Steam Community Market" in description

def parse_date(date_str):
    """Converts the date from format 'DD MMM, YYYY' to a datetime object"""
    try:
        return datetime.strptime(date_str.strip('"'), '%d %b, %Y')
    except ValueError:
        print(f"Could not parse date: '{date_str}'")
        return datetime.min

def create_expenses_excel(transactions):
    """Creates an Excel file with non-market expenses"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Non-Market Expenses"
    
    # Define styles
    header_font = Font(bold=True)
    header_fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
    
    # Add headers
    headers = ["Date", "Description", "Amount (CHF)"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
    
    # Add data
    row = 2
    total = 0
    for trans in transactions:
        if not is_market_transaction(trans['description']) and trans['amount'] < 0:
            ws.cell(row=row, column=1, value=trans['date'])
            ws.cell(row=row, column=2, value=trans['description'])
            ws.cell(row=row, column=3, value=abs(trans['amount']))
            total += abs(trans['amount'])
            row += 1
    
    # Add total row
    total_row = row
    ws.cell(row=total_row, column=1, value="TOTAL")
    ws.cell(row=total_row, column=2, value="")
    ws.cell(row=total_row, column=3, value=total)
    
    # Style total row
    for col in range(1, 4):
        cell = ws.cell(row=total_row, column=col)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")
    
    # Adjust column widths
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column].width = adjusted_width
    
    # Save the workbook
    wb.save("steam_expenses.xlsx")
    print("\nExpenses have been saved to 'steam_expenses.xlsx'")

def analyze_transactions():
    total_earned = 0.0
    total_spent = 0.0
    market_earned = 0.0
    market_spent = 0.0
    
    # List for all transactions
    transactions = []
    
    try:
        with open("steam_wallet_transactions.csv", "r", encoding='utf-8') as file:
            reader = csv.DictReader(file)
            for row in reader:
                change_amount = row['Change']
                description = row['Beschreibung'].split('\n')[0].strip()  # Only first line of description
                date = row['Datum']
                
                # Debug: Show transaction information
                print(f"\nDescription: {description}")
                print(f"Change: '{change_amount}'")
                print(f"Date: '{date}'")
                
                amount = parse_amount(change_amount)
                print(f"Parsed amount: {amount}")
                
                # Save transaction
                transactions.append({
                    'date': date,
                    'date_obj': parse_date(date),  # For sorting
                    'description': description,
                    'amount': amount,
                    'raw_amount': abs(amount)  # Always positive for sorting
                })
                
                # Track transactions
                if is_market_transaction(description):
                    if amount > 0:
                        market_earned += amount
                        print(f"Market earnings found: {description} - CHF {amount:.2f}")
                    else:
                        market_spent += abs(amount)
                else:
                    if amount < 0:
                        total_spent += abs(amount)
                        print(f"Expense found: {description} - CHF {abs(amount):.2f}")
    
    except Exception as e:
        print(f"Error reading CSV file: {e}")
        return
    
    # Sort transactions by date
    transactions.sort(key=lambda x: x['date_obj'])
    
    # Create Excel file for expenses
    create_expenses_excel(transactions)
    
    # Create analysis text
    analysis = f"""=== Steam Transaction Analysis ===
Period: {transactions[0]['date']} to {transactions[-1]['date']}

=== Market Transactions ===
Market earnings: CHF {market_earned:.2f}
Market expenses: CHF {market_spent:.2f}
Market balance: CHF {(market_earned - market_spent):.2f}

=== Other Transactions ===
Total expenses: CHF {total_spent:.2f}

=== Overall Summary ===
Total earnings (Market only): CHF {market_earned:.2f}
Total expenses (all): CHF {(total_spent + market_spent):.2f}
Net balance: CHF {(market_earned - total_spent - market_spent):.2f}

Top 5 largest transactions (excluding Market):
"""
    
    # Sort for Top 5 by amount
    transactions_by_amount = sorted(transactions, key=lambda x: x['raw_amount'], reverse=True)
    
    # Add Top 5 transactions
    count = 1
    for trans in transactions_by_amount:
        if not is_market_transaction(trans['description']):
            transaction_type = "EXPENSE"
            analysis += f"{count}. [{transaction_type}] {trans['description']}: CHF {abs(trans['amount']):.2f}\n"
            count += 1
            if count > 5:
                break
    
    # Save analysis
    with open("steam_analysis.txt", "w", encoding='utf-8') as file:
        file.write(analysis)
    
    print("\nAnalysis has been saved to 'steam_analysis.txt'")

if __name__ == "__main__":
    analyze_transactions() 